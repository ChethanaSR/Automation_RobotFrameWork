import json
import jsonpath
import openpyxl

objWorkbook = openpyxl.load_workbook("C:\\Users\\rchethana\\Documents\\Practise_RobotFramework\\Testdata\\LoginCredentials.xlsx")


def read_from_json_file(elementname):
    f=open("C:/Users/rchethana/Documents/Practise_RobotFramework/Locators/Element.json", "r")
    resources = json.loads(f.read())
    result = jsonpath.jsonpath(resources,elementname)
    return result[0]

def read_max_rows_from_excel(SheetName):
    objWorkSheet = objWorkbook[SheetName]
    return objWorkSheet.max_row

def Read_Cell_Value(SheetName,row,col):
    objWorkSheet = objWorkbook[SheetName]
    return (objWorkSheet.cell(int(row),int(col)).value)










